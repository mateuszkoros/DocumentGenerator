import os
import sys
import openpyxl
import platform
from docx import Document
from dotenv import load_dotenv


load_dotenv()


def check_environment_variables():
    if (
        'TEMPLATE_FILE' not in os.environ or
        'VARIABLES_FILE' not in os.environ or
        'OUTPUT_DIRECTORY' not in os.environ
    ):
        print('One of required environment variables is missing', file=sys.stderr)
        exit(1)


def get_variables(variables_sheet):
    variables_dictionary = {}

    # get variables list
    header_iterator = 0
    for column in variables_sheet.iter_cols(1, variables_sheet.max_column):
        variables_dictionary[column[0].value] = header_iterator
        header_iterator += 1

    return variables_dictionary


if __name__ == '__main__':
    check_environment_variables()
    os.makedirs(os.getenv('OUTPUT_DIRECTORY'), exist_ok=True)

    workbook = openpyxl.load_workbook(filename=os.getenv('VARIABLES_FILE'))
    sheet = workbook.worksheets[0]
    variables = get_variables(sheet)

    for row in sheet.iter_rows(min_row=2):
        # variables should always include File column
        output_docx = f'{os.getenv('OUTPUT_DIRECTORY')}{row[variables['File']].value}.docx'
        doc = Document(os.getenv('TEMPLATE_FILE'))
        for paragraph in doc.paragraphs:
            for variable in [*variables.keys()]:
                placeholder = f'[[{variable}]]'
                if placeholder in paragraph.text:
                    inline = paragraph.runs
                    for i in range(len(inline)):
                        if placeholder in inline[i].text:
                            text = inline[i].text.replace(placeholder, row[variables[variable]].value)
                            inline[i].text = text
        doc.save(output_docx)
        if platform.system() == 'Linux':
            os.system(f'libreoffice --headless --convert-to pdf '
                      f'--outdir {os.getenv('OUTPUT_DIRECTORY')}'
                      f' {os.getenv('OUTPUT_DIRECTORY')}*.docx')
        else:
            os.system(f'rocketpdf parsedocxs {os.getenv('OUTPUT_DIRECTORY')}')

